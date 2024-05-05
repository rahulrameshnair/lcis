#  excel.py

"""
Module for different types of formatting that can be performed on an excel file (.xlsx) containing a life cycle inventory.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def excel_format(excel_file, output_excel_file):
    """
    Performs basic formatting of the LCI excel file based on keywords

    Args:
    excel_file (str): path of the input excel file
    output_excel_file (str): desired name for the output excel file (e.g. demo.xlsx)
    """
    workbook = openpyxl.load_workbook(excel_file)

    # Specify the worksheet to work with
    worksheet = workbook.active

    # Specify a list of keywords to format and their respective formatting attributes
    word_to_format_list = [
        "Project parameters",
        "Database parameters",
        "Activity",
        "Parameters",
        "Exchanges",
    ]
    word_to_format_list2 = [
        "name",
        "location",
        "database",
        "categories",
        "code",
        "reference product",
        "type",
        "unit",
        "amount",
        "worksheet name",
    ]

    # Modify the required formatting attributes here.
    formatting_attributes = [
        (
            Font(color="fafcfc", size=16.5, bold=True),
            PatternFill(start_color="029cf5", end_color="029cf5", fill_type="solid"),
            Alignment(horizontal="center"),
        ),
        (
            Font(color="fafcfc", size=16.5, bold=True),
            PatternFill(start_color="4605fa", end_color="4605fa", fill_type="solid"),
            Alignment(horizontal="center"),
        ),
        (
            Font(color="fafcfc", size=16, bold=True),
            PatternFill(start_color="1a7802", end_color="1a7802", fill_type="solid"),
            Alignment(horizontal="center"),
        ),
        (
            Font(color="fafcfc", size=14, bold=True),
            PatternFill(start_color="c5c7c7", end_color="c5c7c7", fill_type="solid"),
            Alignment(horizontal="center"),
        ),
        (
            Font(color="fafcfc", size=13, bold=True),
            PatternFill(start_color="050505", end_color="050505", fill_type="solid"),
            Alignment(horizontal="center"),
        ),
    ]

    # Iterate through the rows and columns
    for row in worksheet.iter_rows(
        min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column
    ):
        for cell in row:
            cell_value = str(cell.value)
            for i, word_to_format in enumerate(word_to_format_list):
                if word_to_format in cell_value:
                    font, fill, align = formatting_attributes[i]
                    cell.font = font
                    cell.fill = fill
                    cell.alignment = align
                elif cell_value in word_to_format_list2:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(size=12, italic=True)

    # Set the width for all columns
    for column in worksheet.columns:
        for cell in column:
            column_letter = openpyxl.utils.get_column_letter(cell.column)
            worksheet.column_dimensions[column_letter].width = 55

    # Save the modified Excel file
    workbook.save(output_excel_file)

    print(
        f"Formatting applied, column width set, and Excel file saved as '{output_excel_file}'."
    )
